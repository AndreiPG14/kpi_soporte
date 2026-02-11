import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(page_title="Comparador de Asistencia 2 Archivos", layout="wide", initial_sidebar_state="expanded")

st.title("📊 Comparador de Asistencia entre 2 Archivos")

# Sidebar para cargar archivos
st.sidebar.header("📁 Cargar Archivos")

uploaded_file_1 = st.sidebar.file_uploader("Archivo 1 (período 1)", type=["xlsx", "xls"], key="file1")
uploaded_file_2 = st.sidebar.file_uploader("Archivo 2 (período 2)", type=["xlsx", "xls"], key="file2")

def procesar_archivo(uploaded_file, nombre):
    """Procesa un archivo Excel y retorna dataframe preparado"""
    try:
        df = pd.read_excel(uploaded_file)
        
        # Convertir FECHA a datetime
        df['FECHA'] = pd.to_datetime(df['FECHA'], format='%d/%m/%Y')
        
        # Convertir TOTAL a timedelta para calcular horas
        def parse_time_to_hours(time_str):
            try:
                parts = str(time_str).split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2]) if len(parts) > 2 else 0
                return hours + minutes/60 + seconds/3600
            except:
                return 0
        
        df['HORAS'] = df['TOTAL'].apply(parse_time_to_hours)
        df['ARCHIVO'] = nombre
        
        return df
    except Exception as e:
        st.error(f"Error procesando {nombre}: {str(e)}")
        return None

if uploaded_file_1 is not None and uploaded_file_2 is not None:
    # Procesar ambos archivos
    df1 = procesar_archivo(uploaded_file_1, "Archivo 1")
    df2 = procesar_archivo(uploaded_file_2, "Archivo 2")
    
    if df1 is not None and df2 is not None:
        # Obtener información de fechas
        fecha1_min = df1['FECHA'].min().date()
        fecha1_max = df1['FECHA'].max().date()
        fecha2_min = df2['FECHA'].min().date()
        fecha2_max = df2['FECHA'].max().date()
        
        st.sidebar.subheader("📅 Información de Archivos")
        col1, col2 = st.sidebar.columns(2)
        
        with col1:
            st.write("**Archivo 1:**")
            st.caption(f"Desde: {fecha1_min}")
            st.caption(f"Hasta: {fecha1_max}")
            st.caption(f"Registros: {len(df1):,}")
        
        with col2:
            st.write("**Archivo 2:**")
            st.caption(f"Desde: {fecha2_min}")
            st.caption(f"Hasta: {fecha2_max}")
            st.caption(f"Registros: {len(df2):,}")
        
        # Combinar datos
        df_combined = pd.concat([df1, df2], ignore_index=True)
        
        # Tabs para diferentes vistas
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(
            ["📋 Resumen Comparativo", "👥 Cambios de Personal", "📈 Gráficos Comparativos", 
             "🔍 Detalle Trabajadores", "📊 Análisis Actividades", "👨‍💼 Supervisores", "📥 Detalles Completos"]
        )
        
        with tab1:
            st.subheader("Resumen Comparativo de Archivos")
            
            # Métricas generales
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write("**Archivo 1**")
                st.metric("📝 Registros", f"{len(df1):,}")
                st.metric("👤 Trabajadores", df1['TRABAJADOR'].nunique())
                st.metric("⏱️ Total Horas", f"{df1['HORAS'].sum():.2f}h")
                st.metric("📊 Promedio Horas/Registro", f"{df1['HORAS'].mean():.2f}h")
            
            with col2:
                st.write("**Archivo 2**")
                st.metric("📝 Registros", f"{len(df2):,}")
                st.metric("👤 Trabajadores", df2['TRABAJADOR'].nunique())
                st.metric("⏱️ Total Horas", f"{df2['HORAS'].sum():.2f}h")
                st.metric("📊 Promedio Horas/Registro", f"{df2['HORAS'].mean():.2f}h")
            
            with col3:
                st.write("**Diferencias**")
                diff_registros = len(df2) - len(df1)
                diff_trabajadores = df2['TRABAJADOR'].nunique() - df1['TRABAJADOR'].nunique()
                diff_horas = df2['HORAS'].sum() - df1['HORAS'].sum()
                
                st.metric("📝 Cambio Registros", f"{diff_registros:+,}", 
                         delta=f"{(diff_registros/len(df1)*100 if len(df1) > 0 else 0):+.1f}%")
                st.metric("👤 Cambio Trabajadores", f"{diff_trabajadores:+d}",
                         delta=f"{(diff_trabajadores/df1['TRABAJADOR'].nunique()*100 if df1['TRABAJADOR'].nunique() > 0 else 0):+.1f}%")
                st.metric("⏱️ Cambio Horas", f"{diff_horas:+.2f}h",
                         delta=f"{(diff_horas/df1['HORAS'].sum()*100 if df1['HORAS'].sum() > 0 else 0):+.1f}%")
            
            st.divider()
            
            # Resumen por fecha
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Resumen Archivo 1 por Fecha:**")
                resumen1 = df1.groupby(df1['FECHA'].dt.date).agg({
                    'TRABAJADOR': 'nunique',
                    'HORAS': 'sum',
                    'CÓDIGO EMPLEADO': 'count'
                }).rename(columns={
                    'TRABAJADOR': 'Trabajadores',
                    'HORAS': 'Total Horas',
                    'CÓDIGO EMPLEADO': 'Registros'
                })
                st.dataframe(resumen1, use_container_width=True)
            
            with col2:
                st.write("**Resumen Archivo 2 por Fecha:**")
                resumen2 = df2.groupby(df2['FECHA'].dt.date).agg({
                    'TRABAJADOR': 'nunique',
                    'HORAS': 'sum',
                    'CÓDIGO EMPLEADO': 'count'
                }).rename(columns={
                    'TRABAJADOR': 'Trabajadores',
                    'HORAS': 'Total Horas',
                    'CÓDIGO EMPLEADO': 'Registros'
                })
                st.dataframe(resumen2, use_container_width=True)
        
        with tab2:
            st.subheader("Análisis de Cambios de Personal")
            
            # Trabajadores únicos en cada archivo
            trabajadores_1 = set(df1['TRABAJADOR'].unique())
            trabajadores_2 = set(df2['TRABAJADOR'].unique())
            
            solo_en_1 = trabajadores_1 - trabajadores_2
            solo_en_2 = trabajadores_2 - trabajadores_1
            en_ambos = trabajadores_1 & trabajadores_2
            
            # Métricas
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("👥 Archivo 1", len(trabajadores_1))
            col2.metric("👥 Archivo 2", len(trabajadores_2))
            col3.metric("🔄 En ambos", len(en_ambos))
            col4.metric("⚠️ Cambios", len(solo_en_1) + len(solo_en_2))
            
            st.divider()
            
            # Mostrar cambios
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**🚪 Solo en Archivo 1 ({len(solo_en_1)} trabajadores):**")
                if solo_en_1:
                    df_solo_1 = df1[df1['TRABAJADOR'].isin(solo_en_1)].groupby('TRABAJADOR').agg({
                        'HORAS': 'sum',
                        'FECHA': 'nunique',
                        'CÓDIGO EMPLEADO': 'first'
                    }).rename(columns={
                        'HORAS': 'Total Horas',
                        'FECHA': 'Días',
                        'CÓDIGO EMPLEADO': 'Código'
                    }).sort_values('Total Horas', ascending=False)
                    st.dataframe(df_solo_1, use_container_width=True)
                    
                    csv1 = df_solo_1.to_csv()
                    st.download_button(
                        "⬇️ Descargar (Solo Archivo 1)",
                        csv1,
                        "solo_archivo_1.csv",
                        "text/csv"
                    )
                else:
                    st.info("Sin cambios - todos están en Archivo 2 también")
            
            with col2:
                st.write(f"**🆕 Solo en Archivo 2 ({len(solo_en_2)} trabajadores):**")
                if solo_en_2:
                    df_solo_2 = df2[df2['TRABAJADOR'].isin(solo_en_2)].groupby('TRABAJADOR').agg({
                        'HORAS': 'sum',
                        'FECHA': 'nunique',
                        'CÓDIGO EMPLEADO': 'first'
                    }).rename(columns={
                        'HORAS': 'Total Horas',
                        'FECHA': 'Días',
                        'CÓDIGO EMPLEADO': 'Código'
                    }).sort_values('Total Horas', ascending=False)
                    st.dataframe(df_solo_2, use_container_width=True)
                    
                    csv2 = df_solo_2.to_csv()
                    st.download_button(
                        "⬇️ Descargar (Solo Archivo 2)",
                        csv2,
                        "solo_archivo_2.csv",
                        "text/csv"
                    )
                else:
                    st.info("Sin cambios - todos estaban en Archivo 1 también")
            
            st.divider()
            
            # Trabajadores en ambos - comparativa de horas
            st.subheader("👥 Trabajadores en Ambos Archivos - Cambio de Horas")
            
            if en_ambos:
                horas_1_ambos = df1[df1['TRABAJADOR'].isin(en_ambos)].groupby('TRABAJADOR')['HORAS'].sum()
                horas_2_ambos = df2[df2['TRABAJADOR'].isin(en_ambos)].groupby('TRABAJADOR')['HORAS'].sum()
                
                comparativa = pd.DataFrame({
                    'Archivo 1': horas_1_ambos,
                    'Archivo 2': horas_2_ambos,
                }).fillna(0)
                
                comparativa['Cambio'] = comparativa['Archivo 2'] - comparativa['Archivo 1']
                comparativa['% Cambio'] = (comparativa['Cambio'] / comparativa['Archivo 1'] * 100).replace([np.inf, -np.inf], 0)
                comparativa = comparativa.sort_values('Cambio', ascending=False)
                
                st.dataframe(comparativa, use_container_width=True)
                
                # Gráfico comparativo
                fig = go.Figure(data=[
                    go.Bar(name='Archivo 1', y=comparativa.index, x=comparativa['Archivo 1'], orientation='h'),
                    go.Bar(name='Archivo 2', y=comparativa.index, x=comparativa['Archivo 2'], orientation='h')
                ])
                fig.update_layout(
                    title="Comparativa de Horas - Trabajadores Presentes en Ambos Archivos",
                    xaxis_title="Horas",
                    yaxis_title="Trabajador",
                    barmode='group',
                    height=600
                )
                st.plotly_chart(fig, use_container_width=True)
                
                # Descargar
                csv_comparativa = comparativa.to_csv()
                st.download_button(
                    "⬇️ Descargar comparativa",
                    csv_comparativa,
                    "comparativa_horas.csv",
                    "text/csv"
                )
        
        with tab3:
            st.subheader("Gráficos Comparativos")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Top 10 por horas - Archivo 1
                top1 = df1.groupby('TRABAJADOR')['HORAS'].sum().nlargest(10)
                fig1 = px.bar(
                    x=top1.values,
                    y=top1.index,
                    orientation='h',
                    title="Top 10 Trabajadores - Archivo 1",
                    labels={'x': 'Horas', 'y': 'Trabajador'}
                )
                st.plotly_chart(fig1, use_container_width=True)
            
            with col2:
                # Top 10 por horas - Archivo 2
                top2 = df2.groupby('TRABAJADOR')['HORAS'].sum().nlargest(10)
                fig2 = px.bar(
                    x=top2.values,
                    y=top2.index,
                    orientation='h',
                    title="Top 10 Trabajadores - Archivo 2",
                    labels={'x': 'Horas', 'y': 'Trabajador'}
                )
                st.plotly_chart(fig2, use_container_width=True)
            
            col3, col4 = st.columns(2)
            
            with col3:
                # Distribución por actividad - Archivo 1
                act1 = df1.groupby('ACTIVIDAD')['HORAS'].sum().sort_values(ascending=False)
                fig3 = px.pie(
                    values=act1.values,
                    names=act1.index,
                    title="Distribución por Actividad - Archivo 1"
                )
                st.plotly_chart(fig3, use_container_width=True)
            
            with col4:
                # Distribución por actividad - Archivo 2
                act2 = df2.groupby('ACTIVIDAD')['HORAS'].sum().sort_values(ascending=False)
                fig4 = px.pie(
                    values=act2.values,
                    names=act2.index,
                    title="Distribución por Actividad - Archivo 2"
                )
                st.plotly_chart(fig4, use_container_width=True)
            
            # Gráfico de línea - horas por fecha
            col1, col2 = st.columns(2)
            
            with col1:
                horas_fecha_1 = df1.groupby(df1['FECHA'].dt.date)['HORAS'].sum()
                fig5 = px.line(
                    x=horas_fecha_1.index,
                    y=horas_fecha_1.values,
                    title="Total de Horas por Fecha - Archivo 1",
                    labels={'x': 'Fecha', 'y': 'Horas'}
                )
                st.plotly_chart(fig5, use_container_width=True)
            
            with col2:
                horas_fecha_2 = df2.groupby(df2['FECHA'].dt.date)['HORAS'].sum()
                fig6 = px.line(
                    x=horas_fecha_2.index,
                    y=horas_fecha_2.values,
                    title="Total de Horas por Fecha - Archivo 2",
                    labels={'x': 'Fecha', 'y': 'Horas'}
                )
                st.plotly_chart(fig6, use_container_width=True)
        
        with tab4:
            st.subheader("Búsqueda de Trabajadores Específicos")
            
            # Seleccionar trabajador
            trabajadores_comunes = list(en_ambos)
            
            if trabajadores_comunes:
                trabajador_seleccionado = st.selectbox("Selecciona un trabajador", sorted(trabajadores_comunes))
                
                # Datos del trabajador
                datos_trab_1 = df1[df1['TRABAJADOR'] == trabajador_seleccionado]
                datos_trab_2 = df2[df2['TRABAJADOR'] == trabajador_seleccionado]
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write("**Archivo 1**")
                    st.metric("Registros", len(datos_trab_1))
                    st.metric("Horas", f"{datos_trab_1['HORAS'].sum():.2f}h")
                    st.metric("Días", datos_trab_1['FECHA'].nunique())
                
                with col2:
                    st.write("**Archivo 2**")
                    st.metric("Registros", len(datos_trab_2))
                    st.metric("Horas", f"{datos_trab_2['HORAS'].sum():.2f}h")
                    st.metric("Días", datos_trab_2['FECHA'].nunique())
                
                with col3:
                    st.write("**Cambio**")
                    diff_horas = datos_trab_2['HORAS'].sum() - datos_trab_1['HORAS'].sum()
                    diff_dias = datos_trab_2['FECHA'].nunique() - datos_trab_1['FECHA'].nunique()
                    st.metric("Cambio Horas", f"{diff_horas:+.2f}h")
                    st.metric("Cambio Días", f"{diff_dias:+d}")
                
                st.divider()
                
                # Detalles
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**Detalles - Archivo 1**")
                    st.dataframe(
                        datos_trab_1[['FECHA', 'ACTIVIDAD', 'H. INICIO', 'H. FIN', 'HORAS', 'GRUPO']].sort_values('FECHA'),
                        use_container_width=True
                    )
                
                with col2:
                    st.write("**Detalles - Archivo 2**")
                    st.dataframe(
                        datos_trab_2[['FECHA', 'ACTIVIDAD', 'H. INICIO', 'H. FIN', 'HORAS', 'GRUPO']].sort_values('FECHA'),
                        use_container_width=True
                    )
            else:
                st.info("No hay trabajadores comunes en ambos archivos")
        
        with tab5:
            st.subheader("Análisis de Actividades")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Actividades - Archivo 1**")
                act_detail_1 = df1.groupby('ACTIVIDAD').agg({
                    'HORAS': 'sum',
                    'TRABAJADOR': 'nunique',
                    'CÓDIGO EMPLEADO': 'count'
                }).rename(columns={
                    'HORAS': 'Total Horas',
                    'TRABAJADOR': 'Trabajadores',
                    'CÓDIGO EMPLEADO': 'Registros'
                }).sort_values('Total Horas', ascending=False)
                st.dataframe(act_detail_1, use_container_width=True)
            
            with col2:
                st.write("**Actividades - Archivo 2**")
                act_detail_2 = df2.groupby('ACTIVIDAD').agg({
                    'HORAS': 'sum',
                    'TRABAJADOR': 'nunique',
                    'CÓDIGO EMPLEADO': 'count'
                }).rename(columns={
                    'HORAS': 'Total Horas',
                    'TRABAJADOR': 'Trabajadores',
                    'CÓDIGO EMPLEADO': 'Registros'
                }).sort_values('Total Horas', ascending=False)
                st.dataframe(act_detail_2, use_container_width=True)
            
            st.divider()
            
            # Actividades nuevas/que desaparecieron
            actividades_1 = set(df1['ACTIVIDAD'].unique())
            actividades_2 = set(df2['ACTIVIDAD'].unique())
            
            col1, col2 = st.columns(2)
            
            with col1:
                nuevas = actividades_2 - actividades_1
                st.write(f"**✨ Actividades nuevas en Archivo 2 ({len(nuevas)}):**")
                if nuevas:
                    for act in sorted(nuevas):
                        horas = df2[df2['ACTIVIDAD'] == act]['HORAS'].sum()
                        st.text(f"  • {act} ({horas:.2f}h)")
                else:
                    st.info("Sin actividades nuevas")
            
            with col2:
                desaparecidas = actividades_1 - actividades_2
                st.write(f"**❌ Actividades que desaparecieron ({len(desaparecidas)}):**")
                if desaparecidas:
                    for act in sorted(desaparecidas):
                        horas = df1[df1['ACTIVIDAD'] == act]['HORAS'].sum()
                        st.text(f"  • {act} ({horas:.2f}h)")
                else:
                    st.info("Sin actividades desaparecidas")
        
        with tab6:
            st.subheader("Análisis de Supervisores - Comparativa entre Archivos")
            
            # Comparativa de supervisores lado a lado
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**ARCHIVO 1 - Supervisores y Grupos**")
                supervisores_1 = df1.groupby(['SUPERVISOR', 'GRUPO']).agg({
                    'TRABAJADOR': 'nunique',
                    'HORAS': 'sum',
                    'FECHA': 'nunique',
                    'CÓDIGO EMPLEADO': 'count'
                }).rename(columns={
                    'TRABAJADOR': 'Trabajadores',
                    'HORAS': 'Total Horas',
                    'FECHA': 'Días',
                    'CÓDIGO EMPLEADO': 'Registros'
                }).sort_values('Total Horas', ascending=False)
                st.dataframe(supervisores_1, use_container_width=True)
            
            with col2:
                st.write("**ARCHIVO 2 - Supervisores y Grupos**")
                supervisores_2 = df2.groupby(['SUPERVISOR', 'GRUPO']).agg({
                    'TRABAJADOR': 'nunique',
                    'HORAS': 'sum',
                    'FECHA': 'nunique',
                    'CÓDIGO EMPLEADO': 'count'
                }).rename(columns={
                    'TRABAJADOR': 'Trabajadores',
                    'HORAS': 'Total Horas',
                    'FECHA': 'Días',
                    'CÓDIGO EMPLEADO': 'Registros'
                }).sort_values('Total Horas', ascending=False)
                st.dataframe(supervisores_2, use_container_width=True)
            
            st.divider()
            
            # Cambios en supervisores
            st.subheader("📊 Cambios de Supervisores entre Archivos")
            
            supervisores_set_1 = set(df1['SUPERVISOR'].unique())
            supervisores_set_2 = set(df2['SUPERVISOR'].unique())
            
            solo_en_1 = supervisores_set_1 - supervisores_set_2
            solo_en_2 = supervisores_set_2 - supervisores_set_1
            en_ambos = supervisores_set_1 & supervisores_set_2
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("👨‍💼 Archivo 1", len(supervisores_set_1))
            col2.metric("👨‍💼 Archivo 2", len(supervisores_set_2))
            col3.metric("🔄 En ambos", len(en_ambos))
            col4.metric("⚠️ Cambios", len(solo_en_1) + len(solo_en_2))
            
            st.divider()
            
            # Supervisores nuevos y que desaparecieron
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**🚪 Solo en Archivo 1 ({len(solo_en_1)}):**")
                if solo_en_1:
                    df_solo_sup_1 = df1[df1['SUPERVISOR'].isin(solo_en_1)].groupby(['SUPERVISOR', 'GRUPO']).agg({
                        'TRABAJADOR': 'nunique',
                        'HORAS': 'sum',
                        'FECHA': 'nunique',
                        'CÓDIGO EMPLEADO': 'count'
                    }).rename(columns={
                        'TRABAJADOR': 'Trabajadores',
                        'HORAS': 'Total Horas',
                        'FECHA': 'Días',
                        'CÓDIGO EMPLEADO': 'Registros'
                    }).sort_values('Total Horas', ascending=False)
                    st.dataframe(df_solo_sup_1, use_container_width=True)
                    
                    csv_solo_1 = df_solo_sup_1.to_csv()
                    st.download_button(
                        "⬇️ Descargar (Solo Archivo 1)",
                        csv_solo_1,
                        "supervisores_solo_archivo_1.csv",
                        "text/csv"
                    )
                else:
                    st.success("✅ Sin supervisores únicos")
            
            with col2:
                st.write(f"**🆕 Solo en Archivo 2 ({len(solo_en_2)}):**")
                if solo_en_2:
                    df_solo_sup_2 = df2[df2['SUPERVISOR'].isin(solo_en_2)].groupby(['SUPERVISOR', 'GRUPO']).agg({
                        'TRABAJADOR': 'nunique',
                        'HORAS': 'sum',
                        'FECHA': 'nunique',
                        'CÓDIGO EMPLEADO': 'count'
                    }).rename(columns={
                        'TRABAJADOR': 'Trabajadores',
                        'HORAS': 'Total Horas',
                        'FECHA': 'Días',
                        'CÓDIGO EMPLEADO': 'Registros'
                    }).sort_values('Total Horas', ascending=False)
                    st.dataframe(df_solo_sup_2, use_container_width=True)
                    
                    csv_solo_2 = df_solo_sup_2.to_csv()
                    st.download_button(
                        "⬇️ Descargar (Solo Archivo 2)",
                        csv_solo_2,
                        "supervisores_solo_archivo_2.csv",
                        "text/csv"
                    )
                else:
                    st.success("✅ Sin supervisores nuevos")
            
            st.divider()
            
            # Comparativa de supervisores comunes
            st.subheader("👨‍💼 Supervisores Comunes - Cambio de Horas y Grupos")
            
            if en_ambos:
                # Crear tabla con grupo incluido
                sup_grupo_1 = df1.groupby('SUPERVISOR').agg({
                    'GRUPO': lambda x: ', '.join(x.unique()),
                    'HORAS': 'sum'
                }).rename(columns={'GRUPO': 'Grupos Archivo 1', 'HORAS': 'Horas Archivo 1'})
                
                sup_grupo_2 = df2.groupby('SUPERVISOR').agg({
                    'GRUPO': lambda x: ', '.join(x.unique()),
                    'HORAS': 'sum'
                }).rename(columns={'GRUPO': 'Grupos Archivo 2', 'HORAS': 'Horas Archivo 2'})
                
                comparativa_sup_grupos = pd.DataFrame({
                    'Grupos Archivo 1': sup_grupo_1['Grupos Archivo 1'],
                    'Horas Archivo 1': sup_grupo_1['Horas Archivo 1'].astype(float),
                    'Grupos Archivo 2': sup_grupo_2['Grupos Archivo 2'],
                    'Horas Archivo 2': sup_grupo_2['Horas Archivo 2'].astype(float),
                }).fillna(0)
                
                comparativa_sup_grupos['Cambio Horas'] = comparativa_sup_grupos['Horas Archivo 2'].astype(float) - comparativa_sup_grupos['Horas Archivo 1'].astype(float)
                comparativa_sup_grupos['% Cambio'] = ((comparativa_sup_grupos['Cambio Horas'] / comparativa_sup_grupos['Horas Archivo 1'].astype(float)) * 100).replace([np.inf, -np.inf], 0)
                comparativa_sup_grupos = comparativa_sup_grupos.sort_values('Cambio Horas', ascending=False)
                
                st.dataframe(comparativa_sup_grupos, use_container_width=True)
                
                # Gráfico comparativo de supervisores
                horas_sup_1 = df1.groupby('SUPERVISOR')['HORAS'].sum()
                horas_sup_2 = df2.groupby('SUPERVISOR')['HORAS'].sum()
                
                fig_sup = go.Figure(data=[
                    go.Bar(name='Archivo 1', y=horas_sup_1.index, x=horas_sup_1.values, orientation='h'),
                    go.Bar(name='Archivo 2', y=horas_sup_2.index, x=horas_sup_2.values, orientation='h')
                ])
                fig_sup.update_layout(
                    title="Comparativa de Horas por Supervisor (con Grupos)",
                    xaxis_title="Horas",
                    yaxis_title="Supervisor",
                    barmode='group',
                    height=500
                )
                st.plotly_chart(fig_sup, use_container_width=True)
                
                # Descargar comparativa
                csv_sup = comparativa_sup_grupos.to_csv()
                st.download_button(
                    "⬇️ Descargar comparativa de supervisores con grupos",
                    csv_sup,
                    "comparativa_supervisores_grupos.csv",
                    "text/csv"
                )
            
            st.divider()
            
            # EXPORTACIÓN CONSOLIDADA DE CAMBIOS
            st.subheader("📥 Exportar Resumen Completo de Cambios")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("📊 Generar Reporte Completo de Cambios", key="generar_reporte_sup"):
                    # Crear un documento consolidado
                    from io import StringIO
                    
                    reporte = StringIO()
                    reporte.write("COMPARATIVA DE SUPERVISORES - RESUMEN DE CAMBIOS\n")
                    reporte.write("=" * 80 + "\n\n")
                    
                    # Resumen general
                    reporte.write("RESUMEN GENERAL\n")
                    reporte.write("-" * 80 + "\n")
                    reporte.write(f"Supervisores Archivo 1: {len(supervisores_set_1)}\n")
                    reporte.write(f"Supervisores Archivo 2: {len(supervisores_set_2)}\n")
                    reporte.write(f"Supervisores en común: {len(en_ambos)}\n")
                    reporte.write(f"Supervisores que desaparecieron: {len(solo_en_1)}\n")
                    reporte.write(f"Supervisores nuevos: {len(solo_en_2)}\n\n")
                    
                    # Solo en Archivo 1
                    reporte.write("SUPERVISORES QUE DESAPARECIERON (Solo en Archivo 1)\n")
                    reporte.write("-" * 80 + "\n")
                    if solo_en_1:
                        df_solo_sup_1 = df1[df1['SUPERVISOR'].isin(solo_en_1)].groupby(['SUPERVISOR', 'GRUPO']).agg({
                            'TRABAJADOR': 'nunique',
                            'HORAS': 'sum',
                            'FECHA': 'nunique',
                            'CÓDIGO EMPLEADO': 'count'
                        }).rename(columns={
                            'TRABAJADOR': 'Trabajadores',
                            'HORAS': 'Total Horas',
                            'FECHA': 'Días',
                            'CÓDIGO EMPLEADO': 'Registros'
                        }).sort_values('Total Horas', ascending=False)
                        reporte.write(df_solo_sup_1.to_string())
                    else:
                        reporte.write("Sin supervisores únicos\n")
                    reporte.write("\n\n")
                    
                    # Solo en Archivo 2
                    reporte.write("SUPERVISORES NUEVOS (Solo en Archivo 2)\n")
                    reporte.write("-" * 80 + "\n")
                    if solo_en_2:
                        df_solo_sup_2 = df2[df2['SUPERVISOR'].isin(solo_en_2)].groupby(['SUPERVISOR', 'GRUPO']).agg({
                            'TRABAJADOR': 'nunique',
                            'HORAS': 'sum',
                            'FECHA': 'nunique',
                            'CÓDIGO EMPLEADO': 'count'
                        }).rename(columns={
                            'TRABAJADOR': 'Trabajadores',
                            'HORAS': 'Total Horas',
                            'FECHA': 'Días',
                            'CÓDIGO EMPLEADO': 'Registros'
                        }).sort_values('Total Horas', ascending=False)
                        reporte.write(df_solo_sup_2.to_string())
                    else:
                        reporte.write("Sin supervisores nuevos\n")
                    reporte.write("\n\n")
                    
                    # Cambios en supervisores comunes
                    reporte.write("CAMBIOS EN SUPERVISORES COMUNES (Con Grupos)\n")
                    reporte.write("-" * 80 + "\n")
                    if en_ambos:
                        sup_grupo_1 = df1.groupby('SUPERVISOR').agg({
                            'GRUPO': lambda x: ', '.join(x.unique()),
                            'HORAS': 'sum'
                        }).rename(columns={'GRUPO': 'Grupos Archivo 1', 'HORAS': 'Horas Archivo 1'})
                        
                        sup_grupo_2 = df2.groupby('SUPERVISOR').agg({
                            'GRUPO': lambda x: ', '.join(x.unique()),
                            'HORAS': 'sum'
                        }).rename(columns={'GRUPO': 'Grupos Archivo 2', 'HORAS': 'Horas Archivo 2'})
                        
                        comparativa_sup_grupos = pd.DataFrame({
                            'Grupos Archivo 1': sup_grupo_1['Grupos Archivo 1'],
                            'Horas Archivo 1': sup_grupo_1['Horas Archivo 1'],
                            'Grupos Archivo 2': sup_grupo_2['Grupos Archivo 2'],
                            'Horas Archivo 2': sup_grupo_2['Horas Archivo 2'],
                        }).fillna('N/A')
                        
                        comparativa_sup_grupos['Cambio Horas'] = comparativa_sup_grupos['Horas Archivo 2'] - comparativa_sup_grupos['Horas Archivo 1']
                        comparativa_sup_grupos = comparativa_sup_grupos.sort_values('Cambio Horas', ascending=False)
                        
                        reporte.write(comparativa_sup_grupos.to_string())
                    else:
                        reporte.write("Sin supervisores en común\n")
                    reporte.write("\n")
                    
                    reporte_texto = reporte.getvalue()
                    st.download_button(
                        "📄 Descargar Reporte en Texto",
                        reporte_texto,
                        "reporte_cambios_supervisores.txt",
                        "text/plain"
                    )
            
            with col2:
                if st.button("📑 Generar Excel Consolidado", key="generar_excel_sup"):
                    from io import BytesIO
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment
                        
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Cambios Supervisores"
                        
                        # Título
                        ws['A1'] = "COMPARATIVA DE SUPERVISORES - CAMBIOS ENTRE ARCHIVOS"
                        ws['A1'].font = Font(bold=True, size=14)
                        ws.merge_cells('A1:F1')
                        
                        row = 3
                        
                        # Solo en Archivo 1
                        ws[f'A{row}'] = "SUPERVISORES QUE DESAPARECIERON (Archivo 1)"
                        ws[f'A{row}'].font = Font(bold=True, size=11)
                        row += 1
                        
                        if solo_en_1:
                            df_solo_sup_1 = df1[df1['SUPERVISOR'].isin(solo_en_1)].groupby(['SUPERVISOR', 'GRUPO']).agg({
                                'TRABAJADOR': 'nunique',
                                'HORAS': 'sum',
                                'FECHA': 'nunique',
                                'CÓDIGO EMPLEADO': 'count'
                            }).rename(columns={
                                'TRABAJADOR': 'Trabajadores',
                                'HORAS': 'Total Horas',
                                'FECHA': 'Días',
                                'CÓDIGO EMPLEADO': 'Registros'
                            }).sort_values('Total Horas', ascending=False)
                            
                            for col_idx, col_name in enumerate(['Supervisor', 'Grupo', 'Trabajadores', 'Total Horas', 'Días', 'Registros'], 1):
                                ws.cell(row, col_idx, col_name)
                            row += 1
                            
                            for (sup, grupo), data in df_solo_sup_1.iterrows():
                                ws.cell(row, 1, sup)
                                ws.cell(row, 2, grupo)
                                ws.cell(row, 3, data['Trabajadores'])
                                ws.cell(row, 4, data['Total Horas'])
                                ws.cell(row, 5, data['Días'])
                                ws.cell(row, 6, data['Registros'])
                                row += 1
                        
                        row += 2
                        
                        # Solo en Archivo 2
                        ws[f'A{row}'] = "SUPERVISORES NUEVOS (Archivo 2)"
                        ws[f'A{row}'].font = Font(bold=True, size=11)
                        row += 1
                        
                        if solo_en_2:
                            df_solo_sup_2 = df2[df2['SUPERVISOR'].isin(solo_en_2)].groupby(['SUPERVISOR', 'GRUPO']).agg({
                                'TRABAJADOR': 'nunique',
                                'HORAS': 'sum',
                                'FECHA': 'nunique',
                                'CÓDIGO EMPLEADO': 'count'
                            }).rename(columns={
                                'TRABAJADOR': 'Trabajadores',
                                'HORAS': 'Total Horas',
                                'FECHA': 'Días',
                                'CÓDIGO EMPLEADO': 'Registros'
                            }).sort_values('Total Horas', ascending=False)
                            
                            for col_idx, col_name in enumerate(['Supervisor', 'Grupo', 'Trabajadores', 'Total Horas', 'Días', 'Registros'], 1):
                                ws.cell(row, col_idx, col_name)
                            row += 1
                            
                            for (sup, grupo), data in df_solo_sup_2.iterrows():
                                ws.cell(row, 1, sup)
                                ws.cell(row, 2, grupo)
                                ws.cell(row, 3, data['Trabajadores'])
                                ws.cell(row, 4, data['Total Horas'])
                                ws.cell(row, 5, data['Días'])
                                ws.cell(row, 6, data['Registros'])
                                row += 1
                        
                        # Convertir a bytes
                        buffer = BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        
                        st.download_button(
                            "📊 Descargar Excel Consolidado",
                            buffer.getvalue(),
                            "cambios_supervisores.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"Error generando Excel: {str(e)}")
            
            st.divider()
            
            # Análisis por fecha y supervisor
            st.subheader("📅 Quién falta tarear por Fecha - Supervisores")
            
            col1, col2 = st.columns(2)
            
            with col1:
                archivo_fecha = st.radio("¿Cuál archivo?", ["Archivo 1", "Archivo 2"], horizontal=True, key="archivo_supervisores")
            
            with col2:
                if archivo_fecha == "Archivo 1":
                    df_analizar = df1.copy()
                    fecha_min_sup = df1['FECHA'].min().date()
                    fecha_max_sup = df1['FECHA'].max().date()
                else:
                    df_analizar = df2.copy()
                    fecha_min_sup = df2['FECHA'].min().date()
                    fecha_max_sup = df2['FECHA'].max().date()
                
                fecha_sup = st.date_input(
                    "Selecciona fecha",
                    fecha_min_sup,
                    min_value=fecha_min_sup,
                    max_value=fecha_max_sup,
                    key="fecha_sup"
                )
            
            # Datos de esa fecha
            df_fecha_sup = df_analizar[df_analizar['FECHA'].dt.date == fecha_sup]
            
            if len(df_fecha_sup) > 0:
                # Supervisores que tarearon
                supervisores_tarearon = set(df_fecha_sup['SUPERVISOR'].unique())
                todos_supervisores = set(df_analizar['SUPERVISOR'].unique())
                supervisores_faltaron = todos_supervisores - supervisores_tarearon
                
                col1, col2, col3 = st.columns(3)
                col1.metric("✅ Tarearon", len(supervisores_tarearon))
                col2.metric("❌ Faltaron", len(supervisores_faltaron))
                col3.metric("📊 Total", len(todos_supervisores))
                
                st.divider()
                
                # Mostrar detalles
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**✅ Supervisores que Tarearon ({len(supervisores_tarearon)}):**")
                    df_tarearon = df_fecha_sup.groupby('SUPERVISOR').agg({
                        'TRABAJADOR': 'nunique',
                        'HORAS': 'sum',
                        'CÓDIGO EMPLEADO': 'count'
                    }).rename(columns={
                        'TRABAJADOR': 'Trabajadores',
                        'HORAS': 'Horas',
                        'CÓDIGO EMPLEADO': 'Registros'
                    }).sort_values('Horas', ascending=False)
                    st.dataframe(df_tarearon, use_container_width=True)
                
                with col2:
                    st.write(f"**❌ Supervisores que Faltaron ({len(supervisores_faltaron)}):**")
                    if supervisores_faltaron:
                        for sup in sorted(supervisores_faltaron):
                            st.error(f"🚫 {sup}")
                    else:
                        st.success("✅ Todos los supervisores tarearon")
            else:
                st.info(f"📭 Sin registros para la fecha {fecha_sup}")
        
        with tab7:
            st.subheader("Detalles Completos de Registros")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                filtro_archivo = st.selectbox("¿Cuál archivo?", ["Ambos", "Archivo 1", "Archivo 2"])
            
            with col2:
                buscar_trabajador = st.text_input("Buscar trabajador", "")
            
            with col3:
                buscar_actividad = st.text_input("Buscar actividad", "")
            
            # Aplicar filtros
            if filtro_archivo == "Archivo 1":
                df_mostrar = df1.copy()
            elif filtro_archivo == "Archivo 2":
                df_mostrar = df2.copy()
            else:
                df_mostrar = df_combined.copy()
            
            if buscar_trabajador:
                df_mostrar = df_mostrar[df_mostrar['TRABAJADOR'].str.contains(buscar_trabajador, case=False, na=False)]
            
            if buscar_actividad:
                df_mostrar = df_mostrar[df_mostrar['ACTIVIDAD'].str.contains(buscar_actividad, case=False, na=False)]
            
            # Mostrar tabla
            st.dataframe(
                df_mostrar.sort_values('FECHA', ascending=False)[
                    ['ARCHIVO', 'FECHA', 'TRABAJADOR', 'ACTIVIDAD', 'H. INICIO', 'H. FIN', 'HORAS', 'GRUPO']
                ],
                use_container_width=True,
                height=400
            )
            
            st.caption(f"Mostrando {len(df_mostrar):,} registros")
            
            # Descargar
            csv_detalle = df_mostrar.to_csv(index=False)
            st.download_button(
                "⬇️ Descargar registros",
                csv_detalle,
                "detalle_asistencia.csv",
                "text/csv"
            )

else:
    st.info("👆 Sube ambos archivos Excel para comenzar la comparación")
    
    col1, col2 = st.columns(2)
    with col1:
        st.write("**Archivo 1** - Período anterior")
    with col2:
        st.write("**Archivo 2** - Período actual")