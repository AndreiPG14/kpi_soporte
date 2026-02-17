"""
TICKETERA DE SOPORTE - VISTA USUARIO
Con autenticación automática de Streamlit Cloud
Archivos en PostgreSQL - BD COMPARTIDA
"""

import streamlit as st
import pandas as pd
import json
import psycopg2
from datetime import datetime
import uuid
import io

# ============================================================
# CONFIGURACIÓN
# ============================================================

st.set_page_config(
    page_title="Ticketera - Usuario",
    layout="wide",
    initial_sidebar_state="expanded"
)

COLUMNAS_REQUERIDAS = ['DNI', 'NOMBRES Y APELLIDOS', 'ACTIVIDAD', 'SUPER', 'FUNDO', 'OBSERVACIONES']

# ============================================================
# CONEXIÓN A BASE DE DATOS
# ============================================================

def get_db_connection():
    """Conectar a PostgreSQL"""
    try:
        conn = psycopg2.connect(st.secrets["database_url"])
        return conn
    except KeyError:
        st.error("❌ Error: variable 'database_url' no encontrada en Secrets")
        st.info("Agrega tu URL de Supabase en Settings → Secrets con el nombre 'database_url'")
        st.stop()
    except psycopg2.OperationalError as e:
        st.error(f"❌ Error de conexión a la BD: {str(e)}")
        st.stop()
    except Exception as e:
        st.error(f"❌ Error desconocido: {type(e).__name__}: {str(e)}")
        st.stop()

def inicializar_db():
    """Inicializar tabla en PostgreSQL"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tickets (
            id TEXT PRIMARY KEY,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            usuario TEXT NOT NULL,
            estado TEXT NOT NULL,
            fecha_creacion TEXT,
            fecha_actualizacion TEXT,
            cantidad_registros INTEGER,
            comentarios TEXT,
            archivo_binario BYTEA,
            nombre_archivo TEXT
        )
    ''')
    
    conn.commit()
    cursor.close()
    conn.close()

def cargar_tickets():
    """Cargar todos los tickets"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM tickets ORDER BY fecha_creacion DESC')
    columnas = [desc[0] for desc in cursor.description]
    
    tickets = []
    for row in cursor.fetchall():
        ticket = dict(zip(columnas, row))
        ticket['comentarios'] = json.loads(ticket['comentarios'] or '[]')
        tickets.append(ticket)
    
    cursor.close()
    conn.close()
    return tickets

def guardar_ticket(ticket, contenido_archivo=None):
    """Guardar ticket en PostgreSQL"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO tickets 
        (id, titulo, descripcion, usuario, estado, fecha_creacion, fecha_actualizacion, cantidad_registros, comentarios, archivo_binario, nombre_archivo)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        ticket['id'],
        ticket['titulo'],
        ticket['descripcion'],
        ticket['usuario'],
        ticket['estado'],
        ticket['fecha_creacion'],
        ticket['fecha_actualizacion'],
        ticket['cantidad_registros'],
        json.dumps(ticket['comentarios']),
        contenido_archivo,
        ticket.get('nombre_archivo')
    ))
    
    conn.commit()
    cursor.close()
    conn.close()

def agregar_comentario(ticket_id, usuario, comentario):
    """Agregar comentario a un ticket"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT comentarios FROM tickets WHERE id = %s', (ticket_id,))
    resultado = cursor.fetchone()
    comentarios = json.loads(resultado[0] or '[]') if resultado else []
    
    comentarios.append({
        'usuario': usuario,
        'texto': comentario,
        'fecha': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    })
    
    cursor.execute('''
        UPDATE tickets 
        SET comentarios = %s, fecha_actualizacion = %s
        WHERE id = %s
    ''', (json.dumps(comentarios), datetime.now().strftime('%d/%m/%Y %H:%M:%S'), ticket_id))
    
    conn.commit()
    cursor.close()
    conn.close()

def obtener_archivo(ticket_id):
    """Obtener archivo de un ticket"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT archivo_binario, nombre_archivo FROM tickets WHERE id = %s', (ticket_id,))
    resultado = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if resultado:
        archivo_binario, nombre_archivo = resultado
        # Convertir memoryview a bytes si es necesario
        if isinstance(archivo_binario, memoryview):
            archivo_binario = bytes(archivo_binario)
        return archivo_binario, nombre_archivo
    return None, None

# ============================================================
# FUNCIONES DE UTILIDAD
# ============================================================

def obtener_usuario_streamlit():
    """Obtener email del usuario invitado en Streamlit Cloud"""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        ctx = get_script_run_ctx()
        
        if ctx and hasattr(ctx, 'user_info') and ctx.user_info:
            if hasattr(ctx.user_info, 'email'):
                email = ctx.user_info.email
                if email and email != "unknown":
                    return email
    except:
        pass
    
    return "usuario_local"

def validar_formato_excel(df):
    """Validar que el Excel tenga el formato correcto"""
    columnas_faltantes = []
    
    for col in COLUMNAS_REQUERIDAS:
        if col not in df.columns:
            columnas_faltantes.append(col)
    
    if columnas_faltantes:
        return False, f"❌ Columnas faltantes: {', '.join(columnas_faltantes)}"
    
    return True, "✅ Formato válido"

def crear_ticket(titulo, descripcion, usuario, archivo, df_datos):
    """Crear nuevo ticket"""
    ticket_id = str(uuid.uuid4())[:8]
    
    ticket = {
        'id': ticket_id,
        'titulo': titulo,
        'descripcion': descripcion,
        'usuario': usuario,
        'estado': 'Abierto',
        'fecha_creacion': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'fecha_actualizacion': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'nombre_archivo': archivo.name,
        'cantidad_registros': len(df_datos),
        'comentarios': []
    }
    
    contenido = archivo.getbuffer().tobytes()
    guardar_ticket(ticket, contenido)
    
    return ticket

# ============================================================
# VISTA PRINCIPAL - USUARIO
# ============================================================

def main():
    """Función principal"""
    
    inicializar_db()
    username = obtener_usuario_streamlit()
    
    if "ticket_creado" not in st.session_state:
        st.session_state.ticket_creado = False
    
    col1, col2, col3 = st.columns([0.6, 0.2, 0.2])
    
    with col1:
        st.title("🎫 TICKETERA DE SOPORTE")
    
    with col3:
        st.write(f"👤 **{username}**")
    
    st.divider()
    
    tab1, tab2 = st.tabs(["📋 Mis Tickets", "➕ Crear Nuevo Ticket"])
    
    with tab1:
        st.subheader("📋 Mis Tickets de Soporte")
        
        tickets = cargar_tickets()
        mis_tickets = [t for t in tickets if t['usuario'] == username]
        
        if not mis_tickets:
            st.info("📭 No tienes tickets aún. ¡Crea tu primer ticket!")
        else:
            estados_disponibles = list(set([t['estado'] for t in mis_tickets]))
            filtro_estado = st.selectbox("Filtrar por estado:", ["Todos"] + estados_disponibles)
            
            tickets_filtrados = mis_tickets
            if filtro_estado != "Todos":
                tickets_filtrados = [t for t in mis_tickets if t['estado'] == filtro_estado]
            
            for ticket in tickets_filtrados:
                if ticket['estado'] == 'Abierto':
                    color = '🔴'
                elif ticket['estado'] == 'En Progreso':
                    color = '🟡'
                else:
                    color = '🟢'
                
                with st.expander(f"{color} [{ticket['id']}] {ticket['titulo']} - {ticket['estado']}"):
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("ID Ticket", ticket['id'])
                        st.metric("Estado", ticket['estado'])
                    
                    with col2:
                        st.metric("Registros", ticket['cantidad_registros'])
                        st.metric("Creado", ticket['fecha_creacion'])
                    
                    with col3:
                        st.metric("Archivo", ticket['nombre_archivo'][:30])
                        st.metric("Actualizado", ticket['fecha_actualizacion'])
                    
                    st.divider()
                    
                    st.write("**Descripción:**")
                    st.write(ticket['descripcion'])
                    
                    st.divider()
                    
                    archivo_binario, nombre_archivo = obtener_archivo(ticket['id'])
                    if archivo_binario:
                        st.download_button(
                            label=f"📥 Descargar: {nombre_archivo}",
                            data=archivo_binario,
                            file_name=nombre_archivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_{ticket['id']}"
                        )
                    
                    st.divider()
                    
                    st.write("**💬 Comentarios:**")
                    
                    if ticket['comentarios']:
                        for com in ticket['comentarios']:
                            st.write(f"👤 **{com['usuario']}** _{com['fecha']}_")
                            st.write(f"> {com['texto']}")
                    else:
                        st.info("Sin comentarios aún")
                    
                    nuevo_comentario = st.text_area(
                        "Agregar comentario:",
                        key=f"comentario_{ticket['id']}",
                        height=80
                    )
                    
                    if st.button("💬 Enviar comentario", key=f"btn_comentario_{ticket['id']}"):
                        if nuevo_comentario.strip():
                            agregar_comentario(ticket['id'], username, nuevo_comentario)
                            st.success("✅ Comentario agregado")
                            st.rerun()
                        else:
                            st.warning("⚠️ Escribe un comentario")
    
    with tab2:
        if st.session_state.ticket_creado:
            st.success("✅ ¡Ticket creado exitosamente!")
            st.info("Puedes crear otro ticket o ir a 'Mis Tickets' para ver el que acabas de crear.")
            st.session_state.ticket_creado = False
            st.divider()
        
        st.subheader("➕ Crear Nuevo Ticket")
        st.write("Por favor carga un archivo Excel con el formato específico.")
        st.divider()
        
        st.write("**📥 Descargar Formato de Ejemplo:**")
        st.info("Descarga este archivo como referencia.")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            
            headers = COLUMNAS_REQUERIDAS
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(color='FFFFFF', bold=True, size=12)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            ejemplos = [
                ['12345678', 'Juan García López', 'Siembra', 'Supervisor 1', 'Fundo A', 'Trabajo realizado correctamente'],
                ['87654321', 'María Rodríguez Pérez', 'Riego', 'Supervisor 2', 'Fundo B', 'Requiere revisión'],
            ]
            
            for row_idx, ejemplo in enumerate(ejemplos, start=2):
                for col_idx, valor in enumerate(ejemplo, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = valor
            
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Descargar Formato de Ejemplo (Excel)",
                data=excel_buffer.getvalue(),
                file_name="FORMATO_EJEMPLO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )
        except Exception as e:
            st.error(f"Error: {str(e)}")
        
        st.divider()
        
        st.write("**📋 Formato Requerido:**")
        cols_display = st.columns(len(COLUMNAS_REQUERIDAS))
        for idx, col in enumerate(COLUMNAS_REQUERIDAS):
            with cols_display[idx]:
                st.write(f"**{col}**")
        
        st.divider()
        
        with st.form("form_crear_ticket", clear_on_submit=True):
            titulo = st.text_input("📌 Título del ticket:")
            descripcion = st.text_area("📝 Descripción (opcional):", height=100)
            st.divider()
            archivo = st.file_uploader("📎 Archivo Excel:", type=['xlsx', 'xls'])
            st.divider()
            col1, col2 = st.columns([0.7, 0.3])
            
            with col2:
                submit_button = st.form_submit_button("✅ Crear Ticket", width='stretch', type="primary")
            
            if submit_button:
                if not titulo:
                    st.error("❌ El título es obligatorio")
                elif archivo is None:
                    st.error("❌ Debes adjuntar un archivo")
                else:
                    try:
                        df = pd.read_excel(archivo)
                        es_valido, mensaje = validar_formato_excel(df)
                        
                        if not es_valido:
                            st.error(mensaje)
                        else:
                            ticket = crear_ticket(
                                titulo,
                                descripcion if descripcion else "Sin descripción",
                                username,
                                archivo,
                                df
                            )
                            
                            st.session_state.ticket_creado = True
                            st.success(f"✅ Ticket #{ticket['id']} creado")
                            st.balloons()
                            import time
                            time.sleep(1.5)
                            st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    main()